from os import startfile, makedirs, listdir
from shutil import copy
from sqlite3 import connect
from PIL import ImageTk, Image
from datetime import date
from random import sample
from openpyxl import Workbook
from tkinter import messagebox as ms

def run_query(query, parametros=()):
    try:
        with connect('BD/database.db') as conn:
            cursor = conn.cursor()
            resultado = cursor.execute(query, parametros)
            conn.commit()
        return resultado
    except:
        ms.showerror('Error Fatal', 'No se pudo conectar a la base de datos, por favor contacte a soporte.')
    
def datos(usuario):
    datos = run_query('select nombre,usuario from empleados where usuario = ?',(usuario,))
    for (nombre,user) in datos:
                pass
    return nombre

def imagen(usuario):
    datos = run_query('select imagen,usuario from empleados where usuario = ?',(usuario,))
    for (img,user) in datos:
                pass
    return img

# def convertirFecha(fecha):
#     fecha = datetime.strptime(date, "%Y-%m-%d")
#     return fecha

def crearDireccion(dir, id, texto):
    if texto == 'socio':
        nuevoArchivo = f'./socios/{id}'
        makedirs(nuevoArchivo, exist_ok=True)
        copy(dir, nuevoArchivo)
        nombre = listdir(nuevoArchivo)
        nuevoArchivo = nuevoArchivo + '/' + nombre[0]
        return nuevoArchivo
        
    elif texto == 'empleado':
        nuevoArchivo = f'./empleados/{id}'
        makedirs(nuevoArchivo, exist_ok=True)
        copy(dir, nuevoArchivo)
        nombre = listdir(nuevoArchivo)
        nuevoArchivo = nuevoArchivo + '/' + nombre[0]
        return nuevoArchivo

def crarImage(imagen):
    img = Image.open(imagen)
    img = img.resize((208, 118), Image.ANTIALIAS)
    img = ImageTk.PhotoImage(img)
    return img

def crarImage2(imagen):
    img = Image.open(imagen)
    img = img.resize((304, 171), Image.ANTIALIAS)
    # img = img.resize((272, 153), Image.ANTIALIAS)
    img = ImageTk.PhotoImage(img)
    return img

def crearImage(imagen):
    img = Image.open(imagen)
    img = img.resize((300, 100), Image.ANTIALIAS)
    img = ImageTk.PhotoImage(img)
    return img

def forIn(datos):
    for elemento in datos:
        pass
    return elemento

def contraAleatoria():
    min = 'qwertyuiopasdfghjklzxcvbnm'
    may = min.upper()
    num = '1234567890'
    sim = '@#$%&*()?><'
    todos = min + may + num + sim
    al = sample(todos,10)
    cont = ''.join(al)
    return cont

def excel():

    empleados = run_query('select * from empleados')
    ventas = run_query('select * from ventas')
    socios = run_query('select * from socios')
    productos = run_query('select * from productos')


    wb = Workbook()
    tabla_socios = wb.active
    tabla_socios.title='Socios'

    tabla_socios.append(('ID', 'Nombre', 'Direccion', 'Telefono', 'Imagen', 'Fecha Inicio', 'Fecha Final'))

    for datos_socios in socios:
        tabla_socios.append(datos_socios)

    
    tabla_ventas= wb.create_sheet('Ventas')

    tabla_ventas.append(('ID', 'Vendido Por', 'Producto', 'Cantidad', 'Dinero','Fecha'))

    for datos_ventas in ventas:
        tabla_ventas.append(datos_ventas)

    tabla_empleados= wb.create_sheet('Empleados')

    tabla_empleados.append(('Usuario', 'Contraseña', 'Nombre', 'Direccion', 'Telefono','Imagen', 'Fecha Contratacion'))

    for datos_empleados in empleados:
        tabla_empleados.append(datos_empleados)

    tabla_productos= wb.create_sheet('Productos')

    tabla_productos.append(('ID', 'Nombre', 'Cantidad', 'Precio'))

    for datos_productos in productos:
        tabla_productos.append(datos_productos)
    
    contacto = wb.create_sheet('Ayuda')
    contacto.append(('Concactanos para poder restaurar tu base de datos', ''))
    contacto.append(('Email', 'idanielmezaledezma@gmail.com'))
    contacto.append(('Telefono', '7828200884'))
    contacto.append(('Facebook', 'SyntraxDL'))
    contacto.append(('MANTEN LA CALMA TU INFORMACION ESTA A SALVO', ''))


    path = f'Copia de seguridad al {date.today()}.xlsx'
    wb.save(path)
    startfile(path)


    

